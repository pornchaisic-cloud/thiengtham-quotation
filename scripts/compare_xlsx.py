"""Compare reference.xlsx (ต้นฉบับ) vs TT-QN-062-26 (app-generated) cell-by-cell.

Output:
- merged cells
- row heights / col widths
- per-cell value (with formula/value)
- images
- print area
"""
import sys
from pathlib import Path
import openpyxl

REF = Path(r"scripts\ref_qn.xlsx")
APP = Path(r"scripts\app_qn062.xlsx")

def describe_xlsx(path: Path, label: str):
    print(f"\n{'='*80}\n{label}: {path.name}\n{'='*80}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"\n--- Sheet: {ws.title} ---")
        print(f"  Dimensions: {ws.dimensions}")
        print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")
        print(f"  Merged cells ({len(ws.merged_cells.ranges)}):")
        for m in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
            print(f"    {m}")
        # Column widths
        print("  Column widths:")
        for letter, dim in sorted(ws.column_dimensions.items()):
            if dim.width:
                print(f"    {letter}: width={dim.width}")
        # Row heights
        print("  Row heights (first 20):")
        rh_items = [(idx, dim.height) for idx, dim in ws.row_dimensions.items() if dim.height]
        for idx, h in sorted(rh_items)[:20]:
            print(f"    Row {idx}: height={h}")
        # Images
        print(f"  Images: {len(ws._images)}")
        for i, img in enumerate(ws._images):
            print(f"    [{i}] anchor={img.anchor._from.col},{img.anchor._from.row} -> {getattr(img.anchor, 'to', 'n/a')}")
        # Print area
        if ws.print_area:
            print(f"  Print area: {ws.print_area}")
        if ws.page_setup.orientation:
            print(f"  Orientation: {ws.page_setup.orientation}")
        # Cell-by-cell text (first 60 rows)
        print("  Cells:")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False):
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).replace("\n", " ")[:80]
                    print(f"    {cell.coordinate}: {val}")

describe_xlsx(REF, "REFERENCE (ต้นฉบับ)")
describe_xlsx(APP, "APP-GENERATED")