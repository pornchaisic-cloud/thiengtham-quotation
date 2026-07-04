"""Extract image positions from reference.xlsx (TwoCellAnchor details)."""
from pathlib import Path
import openpyxl
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

REF = Path(r"scripts\ref_qn.xlsx")

wb = openpyxl.load_workbook(REF)
ws = wb.active

print(f"Sheet: {ws.title}")
for i, img in enumerate(ws._images):
    print(f"\n[Image {i}]")
    print(f"  Type: {type(img.anchor).__name__}")
    print(f"  _from: col={img.anchor._from.col}, colOff={img.anchor._from.colOff}, row={img.anchor._from.row}, rowOff={img.anchor._from.rowOff}")
    if isinstance(img.anchor, TwoCellAnchor) and img.anchor.to is not None:
        t = img.anchor.to
        print(f"  _to:   col={t.col}, colOff={t.colOff}, row={t.row}, rowOff={t.rowOff}")
    elif isinstance(img.anchor, OneCellAnchor):
        ext = img.anchor.ext
        print(f"  ext:   cx={ext.cx}, cy={ext.cy} (pixels)")
    # Format check
    print(f"  format: {img.format if hasattr(img, 'format') else 'unknown'}")
    # Cell positions
    col_from_letter = openpyxl.utils.get_column_letter(img.anchor._from.col + 1)
    row_from_num = img.anchor._from.row + 1
    print(f"  Top-left cell: {col_from_letter}{row_from_num}")
    if isinstance(img.anchor, TwoCellAnchor) and img.anchor.to is not None:
        col_to_letter = openpyxl.utils.get_column_letter(img.anchor.to.col + 1)
        row_to_num = img.anchor.to.row + 1
        print(f"  Bottom-right cell: {col_to_letter}{row_to_num}")